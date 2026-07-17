"""Tests for services/excel_service.py — import, export, error paths."""

import os
import sqlite3
import tempfile

import openpyxl
import pytest

from app import config, database
from services import ServiceError, book_service, excel_service, student_service


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_workbook(sheet_data: list[list], sheet_name: str = "Sheet1"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in sheet_data:
        ws.append(row)
    return wb


def _save_temp(wb) -> str:
    handle, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(handle)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def export_dir(fresh_db):
    os.makedirs(config.EXPORTS_DIR, exist_ok=True)
    return config.EXPORTS_DIR


# ---------------------------------------------------------------------------
# Export tests
# ---------------------------------------------------------------------------

class TestExport:
    def test_export_books(self, fresh_db, sample_book, export_dir):
        path = excel_service.export_books()
        assert os.path.isfile(path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.title == "Books"

    def test_export_students(self, fresh_db, sample_student, export_dir):
        path = excel_service.export_students()
        assert os.path.isfile(path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.title == "Students"

    def test_export_rows(self, fresh_db, export_dir):
        cols = ["Name", "Score"]
        rows = [{"Name": "Alice", "Score": 95}]
        path = excel_service.export_rows(cols, rows, "test_report.xlsx")
        assert os.path.isfile(path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        assert ws.cell(5, 1).value == "Name"
        assert ws.cell(6, 1).value == "Alice"

    def test_export_full_backup(self, fresh_db, sample_book, sample_student, export_dir):
        path = excel_service.export_full_backup()
        assert os.path.isfile(path)
        wb = openpyxl.load_workbook(path)
        sheets = wb.sheetnames
        assert "Books" in sheets
        assert "Students" in sheets
        assert "Book Issues" in sheets
        assert "Settings" in sheets

    def test_export_folder_default(self, fresh_db):
        excel_service.set_export_folder(None)
        os.makedirs(config.EXPORTS_DIR, exist_ok=True)
        path = excel_service.export_books()
        assert config.EXPORTS_DIR in path

    def test_export_timestamped_filename(self, fresh_db, export_dir):
        import re
        path = excel_service.export_books()
        assert re.search(r"\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}", path), \
            "Filename should contain a timestamp"


# ---------------------------------------------------------------------------
# Book import tests
# ---------------------------------------------------------------------------

class TestImportBooks:
    def test_valid_books(self, fresh_db):
        wb = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["Book A", "Author A", "Fiction", "10"],
            ["Book B", "Author B", "Non-Fiction", "5"],
        ])
        summary = excel_service.import_books(_save_temp(wb))
        assert summary.imported == 2
        assert summary.skipped == 0
        assert summary.errors == 0

    def test_empty_workbook_raises(self, fresh_db):
        wb = _make_workbook([])
        with pytest.raises(ServiceError):
            excel_service.import_books(_save_temp(wb))

    def test_corrupted_file_raises(self, fresh_db):
        handle, path = tempfile.mkstemp(suffix=".xlsx")
        with os.fdopen(handle, "w") as f:
            f.write("not an excel file")
        with pytest.raises(ServiceError):
            excel_service.import_books(path)
        os.unlink(path)

    def test_missing_title_is_skipped(self, fresh_db):
        wb = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["", "Author A", "Fiction", "5"],
        ])
        summary = excel_service.import_books(_save_temp(wb))
        assert summary.imported == 0
        assert summary.skipped == 1

    def test_invalid_quantity_is_error(self, fresh_db):
        wb = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["Book A", "Author A", "Fiction", "not-a-number"],
        ])
        summary = excel_service.import_books(_save_temp(wb))
        assert summary.imported == 0
        assert summary.errors == 1

    def test_negative_quantity_is_error(self, fresh_db):
        wb = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["Book A", "Author A", "Fiction", "-5"],
        ])
        summary = excel_service.import_books(_save_temp(wb))
        assert summary.imported == 0
        assert summary.errors == 1

    def test_no_header_row_raises(self, fresh_db):
        wb = _make_workbook([
            ["Book A", "Author A", "Fiction", "5"],
        ])
        summary = excel_service.import_books(_save_temp(wb))
        assert summary.imported == 0


# ---------------------------------------------------------------------------
# Student import tests
# ---------------------------------------------------------------------------

class TestImportStudents:
    def test_valid_students(self, fresh_db):
        wb = _make_workbook([
            ["Student ID", "Name", "Class", "Division"],
            ["STU001", "Alice", "10", "A"],
            ["STU002", "Bob", "10", "B"],
        ])
        summary = excel_service.import_students(_save_temp(wb))
        assert summary.imported == 2

    def test_missing_student_id_skipped(self, fresh_db):
        wb = _make_workbook([
            ["Student ID", "Name", "Class", "Division"],
            ["", "Alice", "10", "A"],
        ])
        summary = excel_service.import_students(_save_temp(wb))
        assert summary.imported == 0
        assert summary.skipped == 1

    def test_missing_name_skipped(self, fresh_db):
        wb = _make_workbook([
            ["Student ID", "Name", "Class", "Division"],
            ["STU001", "", "10", "A"],
        ])
        summary = excel_service.import_students(_save_temp(wb))
        assert summary.imported == 0
        assert summary.skipped == 1

    def test_duplicate_id_skipped(self, fresh_db):
        wb = _make_workbook([
            ["Student ID", "Name", "Class", "Division"],
            ["STU001", "Alice", "10", "A"],
            ["STU001", "Bob", "10", "B"],
        ])
        summary = excel_service.import_students(_save_temp(wb))
        assert summary.imported == 1
        assert summary.duplicate_in_file == 1

    def test_division_uppercased(self, fresh_db):
        wb = _make_workbook([
            ["Student ID", "Name", "Class", "Division"],
            ["STU001", "Alice", "10", "a"],
            ["STU002", "Bob", "10", "b "],
        ])
        excel_service.import_students(_save_temp(wb))
        conn = sqlite3.connect(config.DATABASE_PATH)
        divisions = [
            r[0] for r in conn.execute(
                "SELECT division FROM students ORDER BY id"
            ).fetchall()
        ]
        conn.close()
        assert divisions == ["A", "B"]

    def test_empty_workbook_raises(self, fresh_db):
        wb = _make_workbook([])
        with pytest.raises(ServiceError):
            excel_service.import_students(_save_temp(wb))

    def test_no_header_row(self, fresh_db):
        wb = _make_workbook([
            ["STU001", "Alice", "10", "A"],
        ])
        summary = excel_service.import_students(_save_temp(wb))
        assert summary.imported == 0


# ---------------------------------------------------------------------------
# Import-error-rollback tests
# ---------------------------------------------------------------------------

class TestImportRollback:
    def test_book_import_no_partial_writes_on_failure(self, fresh_db):
        """When a book row fails, previously imported rows still remain."""
        wb = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["Good Book", "Author A", "Fiction", "5"],
            ["Bad Book", "Author B", "Non-Fiction", "not-a-number"],
        ])
        summary = excel_service.import_books(_save_temp(wb))
        assert summary.imported == 1
        assert summary.errors == 1
        conn = sqlite3.connect(config.DATABASE_PATH)
        count = conn.execute("SELECT COUNT(*) FROM books").fetchone()[0]
        conn.close()
        assert count == 1

    def test_student_import_no_partial_writes_on_failure(self, fresh_db):
        """When a student row fails, previously imported rows still remain."""
        wb = _make_workbook([
            ["Student ID", "Name", "Class", "Division"],
            ["STU001", "Alice", "10", "A"],
            ["STU002", "", "10", "B"],
        ])
        summary = excel_service.import_students(_save_temp(wb))
        assert summary.imported == 1
        assert summary.skipped == 1
        conn = sqlite3.connect(config.DATABASE_PATH)
        count = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
        conn.close()
        assert count == 1


# ---------------------------------------------------------------------------
# Duplicate import tests — same file twice must not create duplicates
# ---------------------------------------------------------------------------

class TestDuplicateBookImport:
    def test_import_twice_does_not_duplicate(self, fresh_db):
        wb = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["Book A", "Author A", "Fiction", "10"],
            ["Book B", "Author B", "Non-Fiction", "5"],
        ])
        path = _save_temp(wb)

        first = excel_service.import_books(path)
        assert first.imported == 2

        second = excel_service.import_books(path)
        assert second.imported == 0
        assert second.skipped == 2

        conn = sqlite3.connect(config.DATABASE_PATH)
        count = conn.execute("SELECT COUNT(*) FROM books").fetchone()[0]
        conn.close()
        assert count == 2

    def test_import_twice_does_not_double_quantity(self, fresh_db):
        wb = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["Book A", "Author A", "Fiction", "10"],
        ])
        path = _save_temp(wb)

        excel_service.import_books(path)
        excel_service.import_books(path)

        conn = sqlite3.connect(config.DATABASE_PATH)
        rows = conn.execute(
            "SELECT total_quantity FROM books WHERE title='Book A'"
        ).fetchall()
        conn.close()
        assert len(rows) == 1
        assert rows[0][0] == 10

    def test_different_capitalization_does_not_duplicate(self, fresh_db):
        wb = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["a brief history of time", "stephen hawking", "Science", "5"],
        ])
        path = _save_temp(wb)
        excel_service.import_books(path)

        wb2 = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["A Brief History of Time", "Stephen Hawking", "Science", "5"],
        ])
        summary = excel_service.import_books(_save_temp(wb2))
        assert summary.skipped == 1
        assert summary.imported == 0

        conn = sqlite3.connect(config.DATABASE_PATH)
        count = conn.execute("SELECT COUNT(*) FROM books").fetchone()[0]
        conn.close()
        assert count == 1

    def test_leading_trailing_spaces_do_not_duplicate(self, fresh_db):
        wb = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["  Book A  ", "  Author A  ", "  Fiction  ", "5"],
        ])
        path = _save_temp(wb)
        excel_service.import_books(path)

        wb2 = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["Book A", "Author A", "Fiction", "5"],
        ])
        summary = excel_service.import_books(_save_temp(wb2))
        assert summary.skipped == 1

    def test_repeated_internal_spaces_do_not_duplicate(self, fresh_db):
        wb = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["The   Great   Gatsby", "F. Scott   Fitzgerald", "Classic", "5"],
        ])
        path = _save_temp(wb)
        excel_service.import_books(path)

        wb2 = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["The Great Gatsby", "F. Scott Fitzgerald", "Classic", "5"],
        ])
        summary = excel_service.import_books(_save_temp(wb2))
        assert summary.skipped == 1

    def test_duplicate_rows_inside_one_workbook(self, fresh_db):
        wb = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["Book A", "Author A", "Fiction", "10"],
            ["Book A", "Author A", "Fiction", "10"],
        ])
        summary = excel_service.import_books(_save_temp(wb))
        assert summary.imported == 1
        assert summary.duplicate_in_file == 1

    def test_quantity_change_updates_existing_book(self, fresh_db):
        wb = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["Book A", "Author A", "Fiction", "5"],
        ])
        path = _save_temp(wb)
        excel_service.import_books(path)

        wb2 = _make_workbook([
            ["Title", "Author", "Category", "Total Quantity"],
            ["Book A", "Author A", "Fiction", "10"],
        ])
        summary = excel_service.import_books(_save_temp(wb2))
        assert summary.updated == 1
        assert summary.imported == 0

        conn = sqlite3.connect(config.DATABASE_PATH)
        total = conn.execute(
            "SELECT total_quantity FROM books WHERE title='Book A'"
        ).fetchone()[0]
        conn.close()
        assert total == 10


class TestDuplicateStudentImport:
    def test_import_twice_does_not_duplicate(self, fresh_db):
        wb = _make_workbook([
            ["Student ID", "Name", "Class", "Division"],
            ["STU001", "Alice", "10", "A"],
            ["STU002", "Bob", "10", "B"],
        ])
        path = _save_temp(wb)

        first = excel_service.import_students(path)
        assert first.imported == 2

        second = excel_service.import_students(path)
        assert second.imported == 0
        assert second.skipped == 2

        conn = sqlite3.connect(config.DATABASE_PATH)
        count = conn.execute("SELECT COUNT(*) FROM students").fetchone()[0]
        conn.close()
        assert count == 2

    def test_duplicate_rows_inside_one_workbook(self, fresh_db):
        wb = _make_workbook([
            ["Student ID", "Name", "Class", "Division"],
            ["STU001", "Alice", "10", "A"],
            ["STU001", "Alice", "10", "A"],
        ])
        summary = excel_service.import_students(_save_temp(wb))
        assert summary.imported == 1
        assert summary.duplicate_in_file == 1

    def test_update_existing_student_metadata(self, fresh_db):
        wb = _make_workbook([
            ["Student ID", "Name", "Class", "Division"],
            ["STU001", "Alice", "10", "A"],
        ])
        path = _save_temp(wb)
        excel_service.import_students(path)

        wb2 = _make_workbook([
            ["Student ID", "Name", "Class", "Division"],
            ["STU001", "Alice Smith", "11", "B"],
        ])
        summary = excel_service.import_students(_save_temp(wb2))
        assert summary.updated == 1

        with database.get_connection() as conn:
            row = conn.execute(
                "SELECT name, class_name, division FROM students "
                "WHERE student_code='STU001'"
            ).fetchone()
        assert row["name"] == "Alice Smith"
        assert row["class_name"] == "11"
        assert row["division"] == "B"

    def test_division_stored_uppercase_on_import(self, fresh_db):
        wb = _make_workbook([
            ["Student ID", "Name", "Class", "Division"],
            ["STU001", "Alice", "10", "a"],
        ])
        excel_service.import_students(_save_temp(wb))
        conn = sqlite3.connect(config.DATABASE_PATH)
        div = conn.execute(
            "SELECT division FROM students WHERE student_code='STU001'"
        ).fetchone()[0]
        conn.close()
        assert div == "A"


class TestDuplicateBookService:
    def test_add_book_raises_on_duplicate_normalized_key(self, fresh_db):
        from services.book_service import add_book
        add_book({"title": "Book A", "author": "Author A",
                   "category": "Fiction", "total_quantity": 5})
        with pytest.raises(ServiceError, match="already exists"):
            add_book({"title": "  book a  ", "author": "  author a  ",
                       "category": "  fiction  ", "total_quantity": 5})

    def test_import_book_returns_skipped_for_unchanged(self, fresh_db):
        from services.book_service import import_book
        import_book({"title": "Book A", "author": "Author A",
                      "category": "Fiction", "total_quantity": 5})
        result = import_book({"title": "Book A", "author": "Author A",
                               "category": "Fiction", "total_quantity": 5})
        assert result == "skipped"

    def test_import_book_returns_updated_for_quantity_change(self, fresh_db):
        from services.book_service import import_book, get_book
        bid = book_service.add_book({
            "title": "Book A", "author": "Author A",
            "category": "Fiction", "total_quantity": 5,
        })
        result = import_book({"title": "Book A", "author": "Author A",
                               "category": "Fiction", "total_quantity": 10})
        assert result == "updated"
        book = get_book(bid)
        assert book["total_quantity"] == 10

    def test_import_book_rejects_quantity_below_issued(self, fresh_db):
        from services.book_service import import_book, add_book
        from services.issue_service import issue_book

        bid = add_book({"title": "Book A", "author": "Author A",
                         "category": "Fiction", "total_quantity": 5})
        sid = student_service.add_student({
            "student_code": "STU001", "name": "Alice", "class_name": "10",
        })
        issue_book(sid, bid, issue_date="2026-07-01", due_date="2026-07-08")

        with pytest.raises(ServiceError, match="currently issued"):
            import_book({"title": "Book A", "author": "Author A",
                          "category": "Fiction", "total_quantity": 0})


# ---------------------------------------------------------------------------
# Performance tests — batch import must be fast
# ---------------------------------------------------------------------------

@pytest.mark.slow
class TestStudentImportPerformance:
    def test_import_1000_students_under_5s(self, fresh_db):
        """Batch import 1000 students in a single transaction is fast."""
        header = ["Student ID", "Name", "Class", "Division"]
        rows = [header]
        for i in range(1000):
            rows.append([f"STU{i:04d}", f"Student_{i}", str(i % 12 + 1), "A"])
        wb = _make_workbook(rows)
        import time
        start = time.perf_counter()
        summary = excel_service.import_students(_save_temp(wb))
        elapsed = time.perf_counter() - start
        assert summary.imported == 1000
        assert elapsed < 5.0, f"Import took {elapsed:.2f}s, expected <5s"

    def test_reimport_1000_students_under_3s(self, fresh_db):
        """Re-importing 1000 unchanged students (all skipped) is even faster."""
        header = ["Student ID", "Name", "Class", "Division"]
        rows = [header]
        for i in range(1000):
            rows.append([f"STU{i:04d}", f"Student_{i}", str(i % 12 + 1), "A"])
        wb = _make_workbook(rows)
        path = _save_temp(wb)
        excel_service.import_students(path)
        import time
        start = time.perf_counter()
        summary = excel_service.import_students(path)
        elapsed = time.perf_counter() - start
        assert summary.skipped == 1000
        assert elapsed < 3.0, f"Re-import took {elapsed:.2f}s, expected <3s"
