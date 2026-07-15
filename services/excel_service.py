"""Excel import/export using openpyxl.

Excel is a support format only: importing books/students, exporting reports,
and full backups. It is never the main data store.
"""

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app import config, database, utils
from services import ServiceError, book_service, student_service

_export_folder = None


def set_export_folder(path: str):
    """Update the export folder at runtime (overrides config.EXPORTS_DIR)."""
    global _export_folder
    _export_folder = path

# Header styling for exported sheets.
_HEADER_FILL = PatternFill("solid", fgColor="2563EB")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14, color="111827")


class ImportSummary:
    """Result of an Excel import operation, shown to the user."""

    def __init__(self):
        self.total = 0
        self.imported = 0
        self.skipped = 0
        self.errors = 0
        self.messages = []

    def add_error(self, row_number, reason):
        self.errors += 1
        self.messages.append(f"Row {row_number}: {reason}")

    def add_skip(self, row_number, reason):
        self.skipped += 1
        self.messages.append(f"Row {row_number}: {reason}")

    def as_text(self) -> str:
        lines = [
            f"Total rows: {self.total}",
            f"Imported rows: {self.imported}",
            f"Skipped rows: {self.skipped}",
            f"Error rows: {self.errors}",
        ]
        if self.messages:
            lines.append("")
            lines.append("Details:")
            lines.extend(self.messages[:25])
            if len(self.messages) > 25:
                lines.append(f"...and {len(self.messages) - 25} more.")
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Header reading helper
# ---------------------------------------------------------------------------
def _header_map(worksheet) -> dict:
    """Map normalized header text -> column index from the first row."""
    headers = {}
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if first_row is None:
        return headers
    for idx, value in enumerate(first_row):
        if value is None:
            continue
        key = str(value).strip().lower()
        headers[key] = idx
    return headers


def _cell(row, headers, *names):
    """Return the first matching cell value for any of the given header names."""
    for name in names:
        idx = headers.get(name.lower())
        if idx is not None and idx < len(row):
            value = row[idx]
            if value is not None and str(value).strip() != "":
                return str(value).strip()
    return ""


# ---------------------------------------------------------------------------
# Import: books
# ---------------------------------------------------------------------------
def import_books(file_path: str) -> ImportSummary:
    """Import books from an Excel file. Returns an ImportSummary.

    Rows missing required fields are skipped. Old Excel files with ISBN
    or Rack Number columns are handled gracefully (those columns are ignored).
    """
    summary = ImportSummary()
    worksheet = _open_first_sheet(file_path)
    headers = _header_map(worksheet)
    if not headers:
        raise ServiceError("The Excel file appears to be empty.")

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        summary.total += 1

        title = _cell(row, headers, "Book Title", "Title")
        total_qty = _cell(row, headers, "Total Quantity", "Total Qty", "Quantity")

        if not title:
            summary.add_skip(row_number, "Missing book title.")
            continue
        try:
            total = int(float(total_qty)) if total_qty else 0
        except (ValueError, TypeError):
            summary.add_error(row_number, "Total quantity is not a number.")
            continue
        if total < 0:
            summary.add_error(row_number, "Total quantity cannot be negative.")
            continue

        available_raw = _cell(row, headers, "Available Quantity", "Available Qty")
        try:
            available = int(float(available_raw)) if available_raw else total
        except (ValueError, TypeError):
            available = total
        available = max(0, min(available, total))

        data = {
            "title": title,
            "author": _cell(row, headers, "Author"),
            "category": _cell(row, headers, "Category"),
            "total_quantity": total,
            "available_quantity": available,
        }

        try:
            book_service.add_book(data)
            summary.imported += 1
        except ServiceError as exc:
            summary.add_error(row_number, str(exc))

    utils.log_activity("BOOKS_IMPORTED",
                       f"Imported {summary.imported} books from Excel")
    return summary


# ---------------------------------------------------------------------------
# Import: students
# ---------------------------------------------------------------------------
def import_students(file_path: str) -> ImportSummary:
    """Import students from Excel. Duplicate Student IDs are skipped."""
    summary = ImportSummary()
    worksheet = _open_first_sheet(file_path)
    headers = _header_map(worksheet)
    if not headers:
        raise ServiceError("The Excel file appears to be empty.")

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        summary.total += 1

        code = _cell(row, headers, "Student ID", "Student Code", "Admission Number")
        name = _cell(row, headers, "Student Name", "Name")

        if not code:
            summary.add_skip(row_number, "Missing Student ID.")
            continue
        if not name:
            summary.add_skip(row_number, "Missing Student Name.")
            continue

        existing = database.fetch_one(
            "SELECT id FROM students WHERE student_code = ?", (code,)
        )
        if existing is not None:
            summary.add_skip(row_number, f"Student ID '{code}' already exists.")
            continue

        data = {
            "student_code": code,
            "name": name,
            "class_name": _cell(row, headers, "Class", "Class Name"),
            "division": _cell(row, headers, "Division"),
        }
        try:
            student_service.add_student(data)
            summary.imported += 1
        except ServiceError as exc:
            summary.add_error(row_number, str(exc))

    utils.log_activity("STUDENTS_IMPORTED",
                       f"Imported {summary.imported} students from Excel")
    return summary


# ---------------------------------------------------------------------------
# Export helpers
# ---------------------------------------------------------------------------
def export_rows(columns, rows, filename: str, sheet_title: str = "Report",
                report_title: str = None) -> str:
    """Write a generic (columns, rows) report to an Excel file in exports/.

    ``columns`` are display headers; ``rows`` are dicts. Values are matched to
    columns by position using each dict's values in insertion order. Returns
    the saved file path.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_sheet_name(sheet_title)

    start_row = 1
    school = utils.get_setting("school_name", "")
    library = utils.get_setting("library_name", "")
    heading = report_title or sheet_title

    # Title + metadata block.
    worksheet.cell(row=1, column=1, value=f"{heading}").font = _TITLE_FONT
    meta = []
    if school:
        meta.append(school)
    if library:
        meta.append(library)
    if meta:
        worksheet.cell(row=2, column=1, value=" — ".join(meta))
    worksheet.cell(row=3, column=1,
                   value=f"Exported on: {utils.now_timestamp()}")
    start_row = 5

    _write_table(worksheet, columns, rows, start_row)

    path = _export_path(filename)
    workbook.save(path)
    return path


def export_books(rows=None) -> str:
    """Export the full books list to Excel."""
    if rows is None:
        rows = book_service.get_books()
    columns = ["Book ID", "Title", "Author", "Category",
               "Total Quantity", "Available Quantity", "Status",
               "Created At", "Updated At"]
    table = [
        {
            "id": b["id"], "title": b["title"], "author": b.get("author"),
            "category": b.get("category"),
            "total_quantity": b["total_quantity"],
            "available_quantity": b["available_quantity"],
            "status": b["status"], "created_at": b.get("created_at"),
            "updated_at": b.get("updated_at"),
        }
        for b in rows
    ]
    return export_rows(columns, table, "books_export.xlsx", "Books", "Books List")


def export_students(rows=None) -> str:
    """Export the full students list to Excel."""
    if rows is None:
        rows = student_service.get_students()
    columns = ["Student ID", "Student Name", "Class", "Division",
               "Status", "Created At", "Updated At"]
    table = [
        {
            "student_code": s["student_code"], "name": s["name"],
            "class_name": s.get("class_name"), "division": s.get("division"),
            "status": s["status"],
            "created_at": s.get("created_at"), "updated_at": s.get("updated_at"),
        }
        for s in rows
    ]
    return export_rows(columns, table, "students_export.xlsx", "Students",
                       "Students List")


def export_full_backup(directory: str = None) -> str:
    """Export books, students, book_issues, and settings to one workbook.

    Saved in the backups folder by default. Returns the saved path.
    """
    workbook = Workbook()

    # Books sheet
    ws_books = workbook.active
    ws_books.title = "Books"
    book_cols = ["Book ID", "Title", "Author", "Category",
                 "Total Quantity", "Available Quantity", "Status",
                 "Created At", "Updated At"]
    _write_table(ws_books, book_cols,
                 [{
                     "id": r["id"], "title": r["title"],
                     "author": r["author"], "category": r["category"],
                     "total_quantity": r["total_quantity"],
                     "available_quantity": r["available_quantity"],
                     "status": r["status"],
                     "created_at": r["created_at"],
                     "updated_at": r["updated_at"],
                 } for r in database.fetch_all("SELECT * FROM books ORDER BY id")],
                 1)

    # Students sheet
    ws_students = workbook.create_sheet("Students")
    student_cols = ["Student ID", "Name", "Class", "Division",
                    "Status", "Created At", "Updated At"]
    _write_table(
        ws_students, student_cols,
        [_student_backup_row(r) for r in database.fetch_all("SELECT * FROM students ORDER BY id")],
        1,
    )

    # Book issues sheet
    ws_issues = workbook.create_sheet("Book Issues")
    issue_cols = ["Issue ID", "Student ID", "Student Name", "Book Title",
                  "Issue Date", "Due Date", "Return Date", "Status", "Remarks"]
    _write_table(ws_issues, issue_cols, _issue_backup_rows(), 1)

    # Settings sheet
    ws_settings = workbook.create_sheet("Settings")
    _write_table(
        ws_settings, ["Setting Key", "Setting Value"],
        [dict(r) for r in database.fetch_all(
            "SELECT setting_key, setting_value FROM settings ORDER BY setting_key")],
        1,
    )

    target_dir = directory or config.BACKUPS_DIR
    os.makedirs(target_dir, exist_ok=True)
    date_part = utils.today_str()
    path = os.path.join(target_dir, f"library_full_backup_{date_part}.xlsx")
    workbook.save(path)
    utils.log_activity("EXCEL_BACKUP", f"Full Excel backup created: {path}")
    return path


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------
def _student_backup_row(row):
    d = dict(row)
    return {
        "student_code": d.get("student_code"), "name": d.get("name"),
        "class_name": d.get("class_name"), "division": d.get("division"),
        "status": d.get("status"),
        "created_at": d.get("created_at"), "updated_at": d.get("updated_at"),
    }


def _issue_backup_rows():
    rows = database.fetch_all(
        "SELECT bi.id AS issue_id, s.student_code, s.name AS student_name, "
        "b.title AS book_title, bi.issue_date, bi.due_date, bi.return_date, "
        "bi.status, bi.remarks FROM book_issues bi "
        "JOIN books b ON bi.book_id = b.id "
        "JOIN students s ON bi.student_id = s.id ORDER BY bi.id"
    )
    return [dict(r) for r in rows]


def _write_table(worksheet, columns, rows, start_row):
    """Write a header row + data rows and auto-size columns."""
    for col_idx, header in enumerate(columns, start=1):
        cell = worksheet.cell(row=start_row, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for r_offset, row in enumerate(rows, start=1):
        values = list(row.values())
        for col_idx in range(len(columns)):
            value = values[col_idx] if col_idx < len(values) else ""
            worksheet.cell(row=start_row + r_offset, column=col_idx + 1,
                           value=value)

    _autosize(worksheet, columns, rows, start_row)


def _autosize(worksheet, columns, rows, start_row):
    for col_idx in range(len(columns)):
        max_len = len(str(columns[col_idx]))
        for row in rows:
            values = list(row.values())
            if col_idx < len(values) and values[col_idx] is not None:
                max_len = max(max_len, len(str(values[col_idx])))
        letter = get_column_letter(col_idx + 1)
        worksheet.column_dimensions[letter].width = min(max_len + 4, 45)


def _open_first_sheet(file_path: str):
    if not file_path or not os.path.exists(file_path):
        raise ServiceError("Selected Excel file could not be found.")
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ServiceError("Invalid Excel file format.") from exc
    return workbook.active


def _safe_sheet_name(name: str) -> str:
    invalid = set('[]:*?/\\')
    clean = "".join(c for c in str(name) if c not in invalid)
    return (clean or "Sheet")[:31]


def _export_path(filename: str) -> str:
    folder = _export_folder or config.EXPORTS_DIR
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, filename)
